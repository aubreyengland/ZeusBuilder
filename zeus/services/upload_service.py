import logging
from pathlib import Path
from zeus import registry
from openpyxl import load_workbook
from collections import OrderedDict
from typing import List, Dict, Optional
from zeus.shared.data_type_models import DataTypeBase
from zeus.exceptions import ZeusCmdError, ZeusWorkbookError

log = logging.getLogger(__name__)


class RowLoadResp:
    """
    Represents an uploaded worksheet row.

    Attributes:
        index (int): Row number in the worksheet
        data: data type model instance if row data is valid
        error (str): Error message if model validation failed
    """
    def __init__(self, index, data=None, error=""):
        self.index: int = index
        self.data: Optional[DataTypeBase] = data
        self.error: str = error


class WorksheetLoadResp:
    """
    Represents an uploaded worksheet from a provisioning workbook.

    Attributes:
        data_type (str): The data type represented by the worksheet rows
        rows (list): RowLoadResp instance for each valued row in the worksheet
        sheet_error (str): Error message if the sheet could not be parsed
    """
    def __init__(self, data_type, rows=None, sheet_error=""):
        self.data_type: str = data_type
        self.rows: List[RowLoadResp] = rows or []
        self.sheet_error: str = sheet_error

    @property
    def error_rows(self):
        return [r for r in self.rows if r.error]

    @property
    def loaded_rows(self):
        return [r for r in self.rows if not r.error]


class UploadSvc:
    """
    Upload Service class.

    The Upload Service is responsible for:
    - Matching worksheets to the tool's data types
    - Getting the custom UploadTask subclass for a data type, if one is registered
    - Creating a base UploadTask instance if a custom class is not registered.
    - Executing the UploadTasks in order based on their priority value
    - Returning a WorksheetLoadResp for each worksheet in the same order
      as the worksheets in the workbook

    The UploadSvc is used by each tool. It does not need to be subclassed or registered.

    Attributes:
        fh: Workbook file handle from form data
        tool (str): five9, zoom, etc.
        wav_files (dict): Holds prompts upload with the workbook (if applicable)
        loaded_worksheets (dict): Worksheet rows parsed into dictionaries keyed by the associated data type
        worksheet_responses (dict): WorksheetLoadResp instances created by the UploadTasks keyed by the data type

    """
    def __init__(self, tool, fh, wav_files=None, **kwargs):
        self.fh = fh
        self._wb = None
        self.tool: str = tool
        self.wav_files: Dict[str, Path] = wav_files or {}
        self.loaded_worksheets: Dict[str, List[dict]] = {}
        self._worksheets_to_load: OrderedDict = OrderedDict()
        self.worksheet_responses: Dict[str, WorksheetLoadResp] = {}

    @property
    def wb(self):
        if not self._wb:
            try:
                self._wb = load_workbook(filename=self.fh, read_only=True, data_only=True)
            except Exception:
                log.exception(f"Workbook Read Failed")
                raise ZeusCmdError(message=f"Cannot read invalid or corrupt xlsx file")

        return self._wb

    @property
    def worksheets_to_load(self) -> dict:
        """
        Match worksheet names to data types to identify worksheets
        that should be loaded.

        Worksheet names should match the data type title from the
        model schema. Allow for case differences and for a worksheet matching
        the data type name.

        Returns:
            (dict): key: data type name, value: worksheet name
        """
        if not self._worksheets_to_load:
            by_data_type = self.upload_data_types
            by_title = {m.schema()["title"].lower(): d for d, m in by_data_type.items()}

            for sheetname in self.wb.sheetnames:
                normalized = sheetname.lower().strip()
                if normalized in by_data_type:
                    data_type = normalized
                else:
                    data_type = by_title.get(normalized)

                if data_type:
                    self._worksheets_to_load[data_type] = sheetname

        return self._worksheets_to_load

    @property
    def upload_data_types(self) -> Dict[str, DataTypeBase]:
        return registry.get_data_types(self.tool, "upload")

    @property
    def ordered_upload_tasks(self) -> List["UploadTask"]:
        """
        Create an UploadTask instance for each uploaded data type
        and return the tasks ordered by task priority.

        Use a custom task in the registry for the data type, if found
        or create an instance of the default UploadTask.
        """
        tasks = []
        for data_type, model_cls in self.upload_data_types.items():
            try:
                task_cls = registry.get_upload_task(self.tool, data_type)
            except LookupError:
                task_cls = UploadTask

            if data_type in self.loaded_worksheets:
                rows = self.loaded_worksheets[data_type]
                task = task_cls(self, model_cls, rows)

                tasks.append(task)

        return sorted(tasks, key=lambda x: x.priority)

    def run(self) -> Dict[str, WorksheetLoadResp]:
        """
        Run the upload tasks for each data type and
        return a dictionary of WorksheetLoadResp instances keyed by data type

        Ensure keys are sorted based on the order of worksheets in the uploaded file.
        """
        self.load_worksheets()

        if not self.loaded_worksheets and not self.worksheet_responses:
            # If not sheets matched an importable data type, raise error to user
            raise ZeusCmdError("No importable worksheets/rows found in the workbook")

        self.run_upload_tasks()

        uploaded_order = list(self.worksheets_to_load)

        in_uploaded_order = sorted(
            self.worksheet_responses.items(),
            key=lambda x: uploaded_order.index(x[0])
        )

        return dict(in_uploaded_order)

    def load_worksheets(self):
        """
        Enumerate sheets in the workbook and load each sheet with a name
        matching a data type.

        The worksheet data is loaded as a list of dictionaries and
        saved in the `self.loaded_worksheets` dict under the data type key.

        If load_worksheet raises and exception, add a WorksheetLoadResp error
        response to the `worksheet_responses` dict.

        Worksheet load order may differ from the order in the workbook so the
        workbook order is saved in the `worksheet_order` list so the
        html tabs can be rendered in the same order as the worksheets in the workbook.
        """
        for data_type, sheetname in self.worksheets_to_load.items():

            try:
                self.loaded_worksheets[data_type] = load_worksheet(self.wb[sheetname])
            except Exception as exc:
                err = getattr(exc, "message", f"Unhandled error: {str(exc)}")
                self.worksheet_responses[data_type] = WorksheetLoadResp(
                    data_type=data_type, sheet_error=err
                )

    def run_upload_tasks(self):
        """
        Run a UploadTask for each loaded worksheet in the ordered
        specified by the UploadTask.priority value.
        """
        for task in self.ordered_upload_tasks:
            task.run()


class UploadTask:
    """
    Upload Task base class.

    Responsible for creating models from the rows in an uploaded worksheet
    and any conversion/preprocessing that might be necessary.

    Creates a RowLoadResp for each row containing the model or validation error.
    Returns a WorksheetLoadResp holding the RowLoadResp's or worksheet parsing error.

    The base UploadTask simply attempts to create a model instance from each row.
    This is sufficient for most data types.

    A custom UploadTask subclass is necessary if:
     - Additional processing is necessary before creating the model instance
     - Data from other worksheets in the workbook must be used in the validation
     - There are uploaded wav files associated with the worksheet

    Custom UploadTasks must be added to the SvcRegistry using the upload_task
    decorator. Ex:

    ```
    @reg.upload_task("tool", "data_type")
    class ToolDataTypeUploadTask(UploadTask):
    ...
    ```

    Attributes:
        svc (UploadSvc): UploadSvc instance that initiated the task
        model_cls (DataTypeBase): data type model type
        rows (list): Dictionaries read from the worksheet rows
    """
    priority = 5

    def __init__(self, svc, model_cls, rows, **kwargs):
        self.svc: UploadSvc = svc
        self.model_cls = model_cls
        self.rows: List[dict] = rows

    @property
    def data_type(self):
        return self.model_cls.schema()["data_type"]

    def run(self) -> None:
        """
        Create a model instance from each row or record the
        validation error.

        Add a WorksheetLoadResp to the UploadSvc.worksheet_responses
        dictionary.
        """
        row_responses = self.validate_rows(self.rows)
        resp = WorksheetLoadResp(data_type=self.data_type, rows=row_responses)
        self.svc.worksheet_responses[self.data_type] = resp
        log.debug(
            f"{self.svc.tool} {self.data_type} upload: {len(resp.loaded_rows)}/"
            f"{len(resp.error_rows)} loaded/failed rows."
        )

    def validate_rows(self, rows):
        row_responses = []

        for idx, row in enumerate(rows, 2):
            resp = self.validate_row(idx, row)
            row_responses.append(resp)

        return row_responses

    def validate_row(self, idx: int, row: dict):
        try:
            model = self.model_cls.from_wb(row)

        except Exception as exc:
            error = getattr(exc, "message", f"Unhandled error: {exc}")
            resp = RowLoadResp(index=idx, error=error)
        else:
            resp = RowLoadResp(index=idx, data=model)

        return resp


def load_worksheet(ws) -> List[dict]:
    """
    Read rows from Excel spreadsheet into dictionaries similar to CSV.DictReader.
    Assumes first row in sheet is column headers and uses these for keys.

    Args:
        ws (Worksheet): Openpyxl worksheet

    Returns:
        processed (list): list of dicts with keys based on column headers
        and values from the corresponding column
    """
    processed = []
    rows = ws.rows

    keys = read_column_headers(next(rows))

    for row in rows:
        values = [
            str(cell.value).strip()
            if cell.value is not None else ""
            for cell in row
        ]

        # drop rows where all cells are empty
        if any([v for v in values]):
            processed.append(dict(zip(keys, values)))

    return processed


def read_column_headers(row):
    """
    Read cell values for header row and return them as a List of strings
    Raise ValueError if blank or duplicate header values are found

    Args:
        row (tuple): Tuple of Cell objects

    Returns:
        keys (list): List of column header values as strings
    """
    headers = []
    for idx, cell in enumerate(row, 1):

        header = str(cell.value).strip()

        if header in headers:
            raise ZeusWorkbookError(message=f"Duplicate header '{header}' in column {idx}")

        headers.append(str(cell.value))

    return headers
