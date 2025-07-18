import logging
import re
from copy import copy
from io import BytesIO
from typing import Optional
from openpyxl import Workbook
from openpyxl.comments import Comment
from tempfile import NamedTemporaryFile
from zeus.exceptions import ZeusCmdError
from dataclasses import dataclass, field
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from zeus.shared.data_type_models import DataTypeFieldDoc, DataTypeBase
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font

log = logging.getLogger(__name__)

HEADER_COLOR = "595959"
RED = "CC0000"
BORDER_COLOR = "FFD9D9D9"
CREATE_COLOR = "597EAA"
UPDATE_COLOR = "8E7CC3"
DELETE_COLOR = "B45F06"
IGNORE_COLOR = "EBEBEB"


DEFAULT_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

HEADER_FONT = Font(
    name="Calibri",
    bold=True,
    italic=False,
    vertAlign=None,
    underline="none",
    strike=False,
    color="FFFFFFFF",
    size="13",
)

DEFAULT_FONT = Font(
    name="Calibri",
    bold=False,
    italic=False,
    vertAlign=None,
    underline="none",
    strike=False,
    size="11",
)


@dataclass
class HeaderFormat:
    font: Optional[Font] = None
    fill: Optional[PatternFill] = None
    border: Optional[Border] = None
    alignment: Optional[Alignment] = None
    comment: Optional[Comment] = None

    def apply(self, cell):
        if self.font:
            cell.font = self.font
        if self.fill:
            cell.fill = self.fill
        if self.border:
            cell.border = self.border
        if self.alignment:
            cell.alignment = self.alignment
        if self.comment:
            cell.comment = self.comment


@dataclass
class ValidationRule:
    """
    Represents an Excel Data Validation rule.
    Holds the information necessary to create and apply
    a data validation rule except the column letter and
    Worksheet, which are provided to the apply command
    at run-time.
    """
    data_validation: DataValidation
    first_row: int = 2
    last_row: int = 1048576

    def apply(self, cd: ColumnDimension):
        col_range = f"{cd.index}{self.first_row}:{cd.index}{self.last_row}"
        cd.parent.add_data_validation(self.data_validation)
        self.data_validation.add(col_range)


@dataclass
class ExpressionRule:
    """
    Represents an Excel Conditional Formatting Expression rule.
    Hold all the information necessary to create and apply a
    conditional formatting rule except the column letter, which is
    provided in the ColumnDimension object passed to the apply method.
    This also provides access to the Worksheet object.
    """
    expression_value: str
    style: DifferentialStyle
    first_row: int = 2
    last_row: int = 1048576
    stop_if_true: bool = True

    def apply(self, cd: ColumnDimension):
        """
        Apply this Conditional formatting expression rule to
        a worksheet column.

        Args:
            cd (ColumnDimension): Openpyxl ColumnDimension instance
        """
        col_letter = cd.index
        col_range = f"{cd.index}{self.first_row}:{cd.index}{self.last_row}"

        rule = Rule(type="expression", dxf=self.style, stopIfTrue=self.stop_if_true)
        rule.formula = [f'${cd.index}{self.first_row}="{self.expression_value}"']

        cd.parent.conditional_formatting.add(col_range, rule)


@dataclass
class ColumnFormat:
    """
    Defines Excel formats to apply to a specific Column
    in a Worksheet.
    """
    font: Optional[Font] = None
    fill: Optional[PatternFill] = None
    border: Optional[Border] = None
    alignment: Optional[Alignment] = None
    resize: bool = True
    width_offset: int = 2
    data_validation: Optional[ValidationRule] = None
    conditional_rules: list[ExpressionRule] = field(default_factory=list)

    def apply(self, cd: ColumnDimension):
        """
        Apply styles to a Column or Cell within a column.
        Styles set on the Column apply to empty cells but are
        not applied to cells with a value. For these the same
        styles must be applied at the cell level.

        Args:
            cd: Openpyxl Cell or ColumnDimension instance
        """
        cell_widths = set()
        self._apply_styles(cd)

        for cell in cd.parent[f"{cd.index}:{cd.index}"]:
            self._apply_styles(cell)
            if cell.value:
                cell_widths.add(len(str(cell.value)))

        for rule in self.conditional_rules:
            rule.apply(cd)

        if self.data_validation:
            self.data_validation.apply(cd)

        if self.resize:
            col_width = max(list(cell_widths)) + self.width_offset
            cd.width = col_width

    def _apply_styles(self, stylable):
        if self.font:
            stylable.font = self.font
        if self.fill:
            stylable.fill = self.fill
        if self.border:
            stylable.border = self.border
        if self.alignment:
            stylable.alignment = self.alignment

    def _get_column_width(self, cd):
        highest_length = 0
        for cell in cd.parent[f"{cd.index}:{cd.index}"]:
            if cell.value:
                dimension = len(str(cell.value))
                if dimension > highest_length:
                    highest_length = dimension

        return highest_length + self.width_offset


class WorkbookCreator:
    """
    Creates an Openpyxl workbook and returns the IO object
    ready for a Flask Response
    """

    def __init__(self, *args, **kwargs):
        self._wb = None

    @property
    def wb(self) -> Workbook:
        if self._wb is None:
            self._wb = Workbook()
            self._wb._sheets = []
        return self._wb

    def run(self, worksheet_data: dict[str, list[dict]]):
        """
        Creates a worksheet for each key in the provided
        workbook_data dictionary.
        """
        try:
            for sheetname, rows in worksheet_data.items():
                self.add_worksheet(sheetname, rows)
                self.format_worksheet(sheetname)

            return self.save_workbook()

        except Exception as exc:
            log.exception("Workbook Creation Failed")
            raise ZeusCmdError("Workbook Creation Failed")

    def add_worksheet(self, sheetname: str, rows: list[dict]):
        """
        Create a worksheet in the provided workbook with the
        provided data and apply defined formatting.

        Args:
            sheetname (str): Worksheet name
            rows (list): List of dicts with keys as column headers and
            values as cell values
        """
        ws = self.wb.create_sheet(sheetname)
        if not rows:
            ws.append(["No Objects Found"])
        else:
            header = list(rows[0].keys())
            ws.append(header)

            for row in rows:
                ws.append([row[h] for h in header])

    def format_worksheet(self, sheetname: str):
        ws = self.wb[sheetname]
        formatter = WsFormatter(ws)
        formatter.format_worksheet()

    def save_workbook(self):
        with NamedTemporaryFile() as temp_wb:
            try:
                if self.wb.worksheets:
                    self.wb.active = self.wb.worksheets[0]

                self.wb.save(temp_wb.name)

            except Exception as exc:
                log.exception("Workbook Creation Failed")
                raise ZeusCmdError(f"Workbook Creation Failed")

            temp_wb.seek(0)
            return BytesIO(temp_wb.read())


class WsFormatter:
    def __init__(self, ws, header_formats=None, column_formats=None):
        self.ws = ws
        self.header_formats: dict[str, HeaderFormat] = header_formats or {}
        self.column_formats: dict[str, ColumnFormat] = column_formats or {}

    @property
    def default_column_formats(self):
        return ColumnFormat(border=DEFAULT_BORDER)

    @property
    def default_header_formats(self):
        red_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        return HeaderFormat(fill=red_fill, font=HEADER_FONT, border=DEFAULT_BORDER)

    def format_worksheet(self):
        self.format_columns()
        self.format_headers()
        self.ws.sheet_view.showGridLines = False

    def format_headers(self):
        """
        Apply font and fill to each header cell
        based on the provided header formats.

        Apply default formatting if formats for a header
        are not provided.
        """
        for cell in self.ws["1:1"]:
            formats = self.header_formats.get(cell.value) or self.default_header_formats
            formats.apply(cell)

    def format_columns(self):
        """
        Formats applied if a ColumnFormat instance
        exists for the column header value.

        Font, border, fill applied to valued cells.
        Validation and conditional rules applied to all
        possible cells.
        """
        for cell in self.ws["1:1"]:
            formats = self.column_formats.get(cell.value) or self.default_column_formats

            col_dimension = self.ws.column_dimensions[cell.column_letter]
            formats.apply(col_dimension)


class ExportWorkbookCreator(WorkbookCreator):
    """
    Creates and applies standard formatting to an Export workbook.
    """
    def __init__(self, data_types, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.data_types: dict = data_types

    def worksheet_model(self, sheetname: str) -> DataTypeBase | None:
        for model in self.data_types.values():
            schema = model.schema()
            if schema["title"] == sheetname:
                return model

        return None

    def add_worksheet(self, sheetname: str, rows: list[dict]):
        """
        Create a worksheet in the provided workbook with the
        provided data and apply defined formatting.

        Args:
            sheetname (str): Worksheet name
            rows (list): List of dicts with keys as column headers and
            values as cell values
        """
        ws = self.wb.create_sheet(sheetname)
        if not rows:
            model = self.worksheet_model(sheetname)
            empty_row = {d.doc_name: "" for d in model.model_doc().doc_fields}
            rows = [empty_row]

        header = list(rows[0].keys())
        ws.append(header)

        for row in rows:
            ws.append([row[h] for h in header])

    def format_worksheet(self, sheetname: str):
        ws = self.wb[sheetname]
        model = self.worksheet_model(sheetname)
        if model:
            builder = ExportWsFormater(model)
            header_formats = builder.header_formats()
            column_formats = builder.column_formats()
        else:
            header_formats, column_formats = None, None

        formatter = WsFormatter(ws, header_formats, column_formats)
        formatter.format_worksheet()


class ExportWsFormater:
    """Build Export worksheet formats based on the model schema."""

    def __init__(self, model):
        self.action_fill_color = "0070C0"
        self.header_fill_color = HEADER_COLOR
        self.model = model

    @property
    def wb_doc_fields(self) -> list[DataTypeFieldDoc]:
        return self.model.model_doc().doc_fields

    def header_formats(self) -> dict:
        """
        Build formats for each header cell that will appear
        in the export workbook.

        Returns:
            (dict): Header value as key, HeaderFormat instance as value
        """
        header_formats = {}

        for doc_field in self.wb_doc_fields:
            color = self.action_fill_color if doc_field.doc_name == "Action" else self.header_fill_color
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            comment = Comment(self.comment_text_for_header(doc_field), "")
            comment.width = 400
            comment.height = 100

            header_formats[doc_field.doc_name] = HeaderFormat(
                fill=fill,
                comment=comment,
                font=HEADER_FONT,
                border=DEFAULT_BORDER,
            )

        return header_formats

    def column_formats(self) -> dict:
        """
        Build formats for worksheet columns.
        Custom data validation is applied to `OneOfStr` field types to provide a dropdown
        of supported options.
        In addition, conditional formatting is applied to the `Action` column

        Returns:
            (dict): Header value as key, ColumnFormat instance as value
        """
        column_formats = {}

        for doc_field in self.wb_doc_fields:

            if doc_field.doc_name == "Action":
                column_formats[doc_field.doc_name] = self.format_for_action_column(doc_field)

            elif doc_field.field_type == "OneOfStr" and doc_field.one_of_values:
                column_formats[doc_field.doc_name] = self.format_for_oneof_column(doc_field)

            else:
                column_formats[doc_field.doc_name] = ColumnFormat(border=DEFAULT_BORDER)

        return column_formats

    @staticmethod
    def comment_text_for_header(doc_field: DataTypeFieldDoc) -> str:
        """
        Build comment text from doc attributes on the model field.
        The comment will include the `doc_required` value (Yes, No, Conditional).

        The `doc_value` and `doc_notes` values are included, if they exist.

        Since these attributes are used for help page generation, they may
        include Markdown formatting.
        Markdown links are stripped but other formatting text may remain
        """

        md_link_rgx = re.compile(r"\[(.+)]\(.+\)")

        def strip_md_link(val):
            out, subbed = md_link_rgx.subn("\g<1>", val)
            if subbed:
                out = f"{out} in Zeus online help"
            return out

        doc_value = strip_md_link(doc_field.doc_value)
        doc_notes = strip_md_link(doc_field.doc_notes)

        comment = f"Value Required: {doc_field.doc_required}\n\n"

        if doc_value:
            comment += f"Supported Values: {doc_value}\n\n"

        if doc_notes:
            comment += f"Note: {doc_notes}\n\n"

        return comment

    @staticmethod
    def format_for_oneof_column(doc_field: DataTypeFieldDoc):
        """
        Build column formats with a validation rule for 'OneOfStr' fields.
        """
        supported_values = doc_field.one_of_values

        validation_formula = ",".join(supported_values)
        dv = DataValidation(type="list", formula1=f'"{validation_formula}"', allow_blank=True)
        validation_rule = ValidationRule(data_validation=dv)

        return ColumnFormat(
            data_validation=validation_rule,
            border=DEFAULT_BORDER,
        )

    def format_for_action_column(self, doc_field: DataTypeFieldDoc):
        """
        Build column formats for the action column. This includes:
        - Data validation for action cells to provide a drop-down
        - Conditional formatting to apply a fill based on the value
        """
        action_font = self.action_font
        ignore_font = self.ignore_font
        supported_actions = doc_field.one_of_values

        validation_formula = ",".join(supported_actions)
        dv = DataValidation(type="list", formula1=f'"{validation_formula}"', allow_blank=True)
        validation_rule = ValidationRule(data_validation=dv)

        rules = []
        for value, color, font in [
            ("CREATE", CREATE_COLOR, action_font),
            ("UPDATE", UPDATE_COLOR, action_font),
            ("DELETE", DELETE_COLOR, action_font),
            ("IGNORE", IGNORE_COLOR, ignore_font),
        ]:
            if value in supported_actions:
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                style = DifferentialStyle(font=font, fill=fill)

                rule = ExpressionRule(expression_value=value, style=style)
                rules.append(rule)

        return ColumnFormat(
            data_validation=validation_rule,
            conditional_rules=rules,
            border=DEFAULT_BORDER,
        )

    @property
    def action_font(self):
        font = copy(DEFAULT_FONT)
        font.bold = True
        font.size = "10"
        font.color = "FFFFFFFF"
        return font

    @property
    def ignore_font(self):
        font = copy(DEFAULT_FONT)
        font.size = "10"
        return font
